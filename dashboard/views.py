from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render, redirect
from main import models
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from openpyxl import load_workbook
from .funcs import pagenator_page, search_with_fields

from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.authtoken.models import Token

def dashboard(request):
    return render(request, 'dashboard/dashboard.html')


# Category
def list_category(request):
    categorys = models.Category.objects.all()
    context = {'categorys': pagenator_page(categorys, 1, request)}
    return render(request, 'category/list.html', context)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def create_category(request):
    if request.method == "POST":
        models.Category.objects.create(
            name = request.POST['name']
        )
        return redirect('dashboard:list_category')
    return render(request, 'category/create.html')


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def detail_category(request, id):
    category = models.Category.objects.get(id=id)
    if request.method == "POST":
        category.name = request.POST['name']
        category.save()
        return redirect('dashboard:list_category')
    return render(request, 'category/detail.html', {'category':category})


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def delete_category(request, id):
    models.Category.objects.get(id=id).delete()
    return redirect('dashboard:list_category')


# Product

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def product_create(request):
    categorys = models.Category.objects.all()
    context = {
        'categorys':categorys
    }
    if request.method == "POST":
        name = request.POST['name']
        description = request.POST['description']
        quantity = request.POST['quantity']
        price = request.POST['price']
        currency = request.POST['currency']
        baner_image = request.FILES['baner_image']
        category_id = request.POST['category_id']
        images = request.FILES.getlist('images')
        product = models.Product.objects.create(
            name=name,
            description = description,
            quantity=quantity,
            price=price,
            currency=currency,
            baner_image=baner_image,
            category_id=category_id
        )
        for image in images:
            models.ProductImage.objects.create(
                image=image,
                product=product
            )
        return redirect('dashboard:dashboard')
    return render(request, 'dashboard/product/create.html', context)


# Auth
@api_view(['POST'])
def sign_up(request):
    user = User.objects.create_user(
                                username=request.data.get('username'),  
                                password=request.data.get('password'),
                                is_staff=True,
                                )
    token = Token.objects.create(user=user)
    return Response({'token':token.key})        


@api_view(["GET"])
def sign_in(request):
    username = request.data.get('username'),  
    password = request.data.get('password'), 
    user = authenticate(username=username, password=password)
    if user.is_authenticated:
        token, _ = Token.objects.get_or_create(user=user)
        return Response({'token': token.key})
    else:    
        return Response({'error': 'Invalid Credentials'})


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def sign_out(request):
    user = request.user
    Token.objects.get(user = user).delete()
    return Response({'result':"Complate"})


@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def user_update(request):
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')
    
    if old_password != new_password:
        user = request.user        
        if user.check_password(old_password):
            user.username = request.data['username']
            user.set_password(new_password)
            user.save()         
            return redirect('dashboard:dashboard')       
    else:
        return Response(status=200)


# Enter product

def product_list(request):
    result = search_with_fields(request)
    print(result)
    items = models.Product.objects.filter(**result)
    print(items)
    context = {'items': pagenator_page(items, 4, request)}
    return render(request, 'dashboard/product/index.html', context)

def prodect_update(request, id):
    product = models.Product.objects.get(id=id)
    category = models.Category.objects.exclude(id=product.category_id)
    if request.method == "POST":
        product.name = request.POST['name']
        product.category.id = request.POST['category_id']
        product.description = request.POST['description']
        product.price = request.POST['price']
        
        product.currency = request.POST['currency']
        models.EnterProduct.objects.create(
            quantity = int(request.POST['quantity_enter']),
            quantity_enter_notation = int(request.POST['quantity_enter_notation']),
            product = product
        )
        if request.POST['quantity_enter_notation'] == "1":
            product.quantity = product.quantity + int(request.POST['quantity_enter'])
        else:
            try:
                product.quantity = product.quantity - int(request.POST['quantity_enter'])
            except:
                product.quantity = 0
        product.save()
        enter = models.EnterProduct.objects.get(is_active=True, product=product)
        enter.is_active = not enter.is_active
        enter.save()
        return redirect('dashboard:product_list')
    context = {
        'product':product,
        'category':category
    }
    return render(request, 'dashboard/product/update.html', context)


def product_detail(request, id):
    product = models.Product.objects.get(id=id)
    categorys = models.Category.objects.all()
    enters = models.EnterProduct.objects.filter(product=product)
    context = {
        'product':product,
        'categorys':categorys,
        'enters':enters
    }
    return render(request, 'dashboard/product/detail.html', context)


def add_card_excel(requset, id):
    queryset = models.EnterProduct.objects.filter(product_id=id)
    data = {
        'Field1': [obj.id for obj in queryset],
        'Field2': [obj.quantity for obj in queryset],
        'Field3': [obj.date.replace(tzinfo=None) for obj in queryset],
    }

    df = pd.DataFrame(data)

    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False, engine='xlsxwriter')

    excel_buffer.seek(0)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=your_data.xlsx'
    response.write(excel_buffer.read())

    return response

def add_all_excel(requset):
    queryset = models.EnterProduct.objects.all()
    data = {
        'Field1': [obj.id for obj in queryset],
        'Field2': [obj.quantity for obj in queryset],
        'Field2': [obj.product for obj in queryset],
        'Field3': [obj.date.replace(tzinfo=None) for obj in queryset],
    }

    df = pd.DataFrame(data)

    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False, engine='xlsxwriter')

    excel_buffer.seek(0)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=your_data.xlsx'
    response.write(excel_buffer.read())

    return response


def excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            quantity, name, date = row
            try:
                product = models.Product.objects.get(name=name)
                models.Product.objects.create(
                    product=product, 
                    quantity=quantity, 
                    date=date
                    )
            except:
                models.Product.objects.create(
                    quantity=quantity, 
                    date=date
                    )
        return redirect('dashboard:excel')
    return render(request, 'dashboard/excel.html') 